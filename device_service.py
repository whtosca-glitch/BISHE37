import base64
import copy
import hashlib
import hmac
import json
import mimetypes
import os
import threading
import uuid
import concurrent.futures
from datetime import datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse
from zipfile import BadZipFile

import pymysql
import requests
from openpyxl import Workbook, load_workbook
from pymysql.cursors import DictCursor


ROOT_DIR = Path(__file__).resolve().parent
DATA_DIR = ROOT_DIR / "data"
SQL_DIR = ROOT_DIR / "sql"
FRONTEND_DIR = ROOT_DIR / "frontend"
FRONTEND_DIST_DIR = FRONTEND_DIR / "dist"
EXCEL_PATH = DATA_DIR / "devices.xlsx"
CONFIG_PATH = ROOT_DIR / "mysql_config.json"
CONFIG_TEMPLATE_PATH = ROOT_DIR / "mysql_config.template.json"
LOGIN_USERNAME = "root"
LOGIN_PASSWORD = "root01"
AUTH_COOKIE_NAME = "device_dashboard_auth"
AUTH_COOKIE_VALUE = "authenticated"


DEFAULT_CONFIG = {
    "server": {
        "host": "127.0.0.1",
        "port": 8000
    },
    "mysql": {
        "enabled": False,
        "host": "YOUR_MYSQL_HOST",
        "port": 3306,
        "user": "YOUR_MYSQL_USER",
        "password": "YOUR_MYSQL_PASSWORD",
        "database": "YOUR_MYSQL_DATABASE",
        "table": "onenet_devices",
        "charset": "utf8mb4",
        "connect_timeout": 30
    }
}


EXCEL_HEADERS = [
    "record_id",
    "display_name",
    "onenet_device_name",
    "product_id",
    "user_id",
    "access_key",
    "auth_version",
    "device_secret",
    "device_id",
    "ip",
    "start_success_count",
    "slot_index",
    "notes",
    "created_at",
    "updated_at"
]

STATUS_HISTORY_SHEET = "status_history"
STATUS_HISTORY_HEADERS = [
    "date",
    "online_count",
    "offline_count",
    "unknown_count",
    "total_devices",
    "updated_at"
]

DEVICE_STATUS_SHEET = "device_status"
DEVICE_STATUS_HEADERS = [
    "record_id",
    "device_name",
    "display_name",
    "status",
    "online",
    "last_seen_at",
    "updated_at"
]


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def deep_merge(base, extra):
    for key, value in extra.items():
        if isinstance(value, dict) and isinstance(base.get(key), dict):
            deep_merge(base[key], value)
        else:
            base[key] = value
    return base


def safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_slot_index(value):
    slot_index = safe_int(value, 0)
    if 1 <= slot_index <= 20:
        return slot_index
    return 0


ALARM_CONFIG_SHEET = "alarm_config"
ALARM_CONFIG_HEADERS = [
    "rule_key",
    "rule_name",
    "threshold_value",
    "alarm_level",
    "enabled",
    "updated_at"
]

ALARM_LOG_SHEET = "alarm_logs"
ALARM_LOG_HEADERS = [
    "log_id",
    "alarm_key",
    "alarm_id",
    "rule_key",
    "rule_name",
    "alarm_time",
    "alarm_status",
    "metric_value",
    "device_record_id",
    "device_name",
    "process_status",
    "processed_at",
    "process_note",
    "process_records",
    "updated_at"
]

ALARM_STATE_SHEET = "alarm_state"
ALARM_STATE_HEADERS = [
    "alarm_key",
    "device_record_id",
    "rule_key",
    "is_active",
    "last_value",
    "updated_at"
]

DEFAULT_ALARM_CONFIG = [
    {"rule_key": "temp", "rule_name": "温度", "threshold_value": 35, "alarm_level": "紧急", "enabled": 1},
    {"rule_key": "humi", "rule_name": "湿度", "threshold_value": 80, "alarm_level": "重要", "enabled": 1},
    {"rule_key": "smoke", "rule_name": "烟雾浓度", "threshold_value": 300, "alarm_level": "次要", "enabled": 1},
    {"rule_key": "status", "rule_name": "在线状态", "threshold_value": 1, "alarm_level": "提醒", "enabled": 1}
]


class DeviceRepository:
    def __init__(self):
        self.lock = threading.RLock()
        DATA_DIR.mkdir(exist_ok=True)
        SQL_DIR.mkdir(exist_ok=True)
        self._ensure_config_files()
        self._ensure_excel_file()

    def _ensure_config_files(self):
        if not CONFIG_TEMPLATE_PATH.exists():
            CONFIG_TEMPLATE_PATH.write_text(
                json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )
        if not CONFIG_PATH.exists():
            CONFIG_PATH.write_text(
                json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )

    def _load_config(self):
        config = copy.deepcopy(DEFAULT_CONFIG)
        if CONFIG_PATH.exists():
            try:
                user_config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
                if isinstance(user_config, dict):
                    deep_merge(config, user_config)
            except json.JSONDecodeError:
                pass
        return config

    def _default_device(self):
        stamp = now_text()
        return {
            "record_id": "default-demo-device",
            "display_name": "demo",
            "onenet_device_name": "demo",
            "product_id": "61ZnL1etk7",
            "user_id": "483694",
            "access_key": "28f571afc3494b249637acada7cc12a7",
            "auth_version": "2020-05-29",
            "device_secret": "Y2Vrdmc5cHZVWGdRb2RMektNME9NNUVKOFlaTHFQMVM=",
            "device_id": "2575390487",
            "ip": "未配置",
            "start_success_count": 1,
            "slot_index": 1,
            "notes": "默认演示设备",
            "created_at": stamp,
            "updated_at": stamp
        }

    def _open_workbook(self):
        workbook = None
        if EXCEL_PATH.exists():
            try:
                workbook = load_workbook(EXCEL_PATH)
            except Exception:
                broken_path = EXCEL_PATH.with_name("devices_broken_" + datetime.now().strftime("%Y%m%d%H%M%S") + ".xlsx")
                try:
                    EXCEL_PATH.replace(broken_path)
                except OSError:
                    pass
        if workbook is None:
            workbook = Workbook()
        self._ensure_sheet(workbook, "devices", EXCEL_HEADERS)
        self._ensure_sheet(workbook, STATUS_HISTORY_SHEET, STATUS_HISTORY_HEADERS)
        self._ensure_sheet(workbook, DEVICE_STATUS_SHEET, DEVICE_STATUS_HEADERS)
        self._ensure_sheet(workbook, ALARM_CONFIG_SHEET, ALARM_CONFIG_HEADERS)
        self._ensure_sheet(workbook, ALARM_LOG_SHEET, ALARM_LOG_HEADERS)
        self._ensure_sheet(workbook, ALARM_STATE_SHEET, ALARM_STATE_HEADERS)
        return workbook

    def _ensure_sheet(self, workbook, sheet_name, headers):
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            if len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet" and workbook.active.max_row == 1 and workbook.active["A1"].value is None:
                sheet = workbook.active
                sheet.title = sheet_name
            else:
                sheet = workbook.create_sheet(title=sheet_name)
        self._ensure_sheet_headers(sheet, headers)
        return sheet

    def _ensure_sheet_headers(self, sheet, headers):
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)

    def _ensure_excel_file(self):
        with self.lock:
            workbook = self._open_workbook()
            self._save_workbook_atomic(workbook)
            workbook.close()
        self._ensure_seed_data()

    def _ensure_seed_data(self):
        records = self._read_excel_records()
        if not records:
            self._write_excel_records([self._default_device()])
        else:
            repaired_records = []
            used_slots = set()
            changed = False
            for record in records:
                next_record = dict(record)
                slot_index = normalize_slot_index(next_record.get("slot_index"))
                if not slot_index or slot_index in used_slots:
                    slot_index = self._find_first_available_slot(repaired_records)
                    if slot_index != normalize_slot_index(next_record.get("slot_index")):
                        changed = True
                if slot_index:
                    used_slots.add(slot_index)
                next_record["slot_index"] = slot_index
                repaired_records.append(next_record)
            if changed:
                self._write_excel_records(repaired_records)
        if not self._read_alarm_config_records():
            self._write_alarm_config_records(DEFAULT_ALARM_CONFIG)

    def _read_excel_records(self):
        records = self._read_sheet_records("devices", EXCEL_HEADERS)
        for record in records:
            record["start_success_count"] = safe_int(record.get("start_success_count"), 0)
            record["slot_index"] = normalize_slot_index(record.get("slot_index"))
        return records

    def _write_excel_records(self, records):
        normalized_records = [self._normalize_record(record, allow_missing=False) for record in records]
        self._write_sheet_records("devices", EXCEL_HEADERS, normalized_records)

    def _read_sheet_records(self, sheet_name, headers):
        with self.lock:
            workbook = self._open_workbook()
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()
        if len(rows) <= 1:
            return []
        records = []
        for row in rows[1:]:
            if not row or not any(cell not in (None, "") for cell in row):
                continue
            record = {}
            for index, header in enumerate(headers):
                value = row[index] if index < len(row) else ""
                record[header] = "" if value is None else value
            records.append(record)
        return records

    def _write_sheet_records(self, sheet_name, headers, records):
        with self.lock:
            workbook = self._open_workbook()
            if sheet_name in workbook.sheetnames:
                old_sheet = workbook[sheet_name]
                workbook.remove(old_sheet)
            sheet = workbook.create_sheet(title=sheet_name)
            self._ensure_sheet_headers(sheet, headers)
            for record in records:
                sheet.append([record.get(header, "") for header in headers])
            self._save_workbook_atomic(workbook)
            workbook.close()

    def _save_workbook_atomic(self, workbook):
        temp_path = EXCEL_PATH.with_name(EXCEL_PATH.stem + ".tmp.xlsx")
        workbook.save(temp_path)
        os.replace(temp_path, EXCEL_PATH)

    def _read_status_history_records(self):
        history_records = self._read_sheet_records(STATUS_HISTORY_SHEET, STATUS_HISTORY_HEADERS)
        for record in history_records:
            record["online_count"] = safe_int(record.get("online_count"), 0)
            record["offline_count"] = safe_int(record.get("offline_count"), 0)
            record["unknown_count"] = safe_int(record.get("unknown_count"), 0)
            record["total_devices"] = safe_int(record.get("total_devices"), 0)
        return history_records

    def _write_status_history_records(self, records):
        normalized = []
        for record in records:
            normalized.append({
                "date": str(record.get("date", "")).strip(),
                "online_count": safe_int(record.get("online_count"), 0),
                "offline_count": safe_int(record.get("offline_count"), 0),
                "unknown_count": safe_int(record.get("unknown_count"), 0),
                "total_devices": safe_int(record.get("total_devices"), 0),
                "updated_at": str(record.get("updated_at", "")).strip() or now_text()
            })
        self._write_sheet_records(STATUS_HISTORY_SHEET, STATUS_HISTORY_HEADERS, normalized)

    def _read_device_status_records(self):
        return self._read_sheet_records(DEVICE_STATUS_SHEET, DEVICE_STATUS_HEADERS)

    def _write_device_status_records(self, records):
        normalized = []
        for record in records:
            normalized.append({
                "record_id": str(record.get("record_id", "")).strip(),
                "device_name": str(record.get("device_name", "")).strip(),
                "display_name": str(record.get("display_name", "")).strip(),
                "status": str(record.get("status", "")).strip(),
                "online": safe_int(record.get("online"), 0),
                "last_seen_at": str(record.get("last_seen_at", "")).strip(),
                "updated_at": str(record.get("updated_at", "")).strip() or now_text()
            })
        self._write_sheet_records(DEVICE_STATUS_SHEET, DEVICE_STATUS_HEADERS, normalized)

    def _read_alarm_config_records(self):
        records = self._read_sheet_records(ALARM_CONFIG_SHEET, ALARM_CONFIG_HEADERS)
        for record in records:
            record["threshold_value"] = safe_float(record.get("threshold_value"), 0)
            record["enabled"] = safe_int(record.get("enabled"), 0)
        return records

    def _write_alarm_config_records(self, records):
        normalized = []
        for record in records:
            normalized.append({
                "rule_key": str(record.get("rule_key", "")).strip(),
                "rule_name": str(record.get("rule_name", "")).strip(),
                "threshold_value": safe_float(record.get("threshold_value"), 0),
                "alarm_level": str(record.get("alarm_level", "")).strip(),
                "enabled": 1 if safe_int(record.get("enabled"), 0) else 0,
                "updated_at": str(record.get("updated_at", "")).strip() or now_text()
            })
        self._write_sheet_records(ALARM_CONFIG_SHEET, ALARM_CONFIG_HEADERS, normalized)

    def _read_alarm_log_records(self):
        records = self._read_sheet_records(ALARM_LOG_SHEET, ALARM_LOG_HEADERS)
        for index, record in enumerate(records):
            if self._is_shifted_alarm_log_record(record):
                record = self._repair_shifted_alarm_log_record(record)
            elif not self._looks_like_log_id(record.get("log_id", "")):
                record["log_id"] = self._build_alarm_log_id(record)
            if not record.get("log_id"):
                record["log_id"] = uuid.uuid4().hex
            process_records = record.get("process_records", "")
            if isinstance(process_records, str) and process_records.strip():
                try:
                    parsed = json.loads(process_records)
                    while isinstance(parsed, str) and parsed.strip():
                        parsed = json.loads(parsed)
                    record["process_records"] = parsed if isinstance(parsed, list) else []
                except Exception:
                    record["process_records"] = []
            elif isinstance(process_records, list):
                record["process_records"] = process_records
            else:
                record["process_records"] = []
            record["process_status"] = str(record.get("process_status", "")).strip() or "未处理"
            record["processed_at"] = str(record.get("processed_at", "")).strip()
            record["process_note"] = str(record.get("process_note", "")).strip()
            records[index] = record
        return records

    def _is_shifted_alarm_log_record(self, record):
        return (
            str(record.get("alarm_time", "")).strip() in ("紧急", "重要", "次要", "提醒")
            and self._looks_like_datetime(record.get("rule_name", ""))
        )

    def _repair_shifted_alarm_log_record(self, record):
        repaired = {
            "log_id": self._build_alarm_log_id({
                "alarm_key": record.get("log_id", ""),
                "alarm_time": record.get("rule_name", ""),
                "alarm_status": record.get("alarm_time", ""),
                "metric_value": record.get("alarm_status", "")
            }),
            "alarm_key": str(record.get("log_id", "")).strip(),
            "alarm_id": str(record.get("alarm_key", "")).strip(),
            "rule_key": str(record.get("alarm_id", "")).strip(),
            "rule_name": str(record.get("rule_key", "")).strip(),
            "alarm_time": str(record.get("rule_name", "")).strip(),
            "alarm_status": str(record.get("alarm_time", "")).strip(),
            "metric_value": str(record.get("alarm_status", "")).strip(),
            "device_record_id": str(record.get("metric_value", "")).strip(),
            "device_name": str(record.get("device_record_id", "")).strip(),
            "process_status": str(record.get("process_status", "")).strip() or "未处理",
            "processed_at": str(record.get("processed_at", "")).strip(),
            "process_note": str(record.get("process_note", "")).strip(),
            "process_records": record.get("process_records", []),
            "updated_at": str(record.get("device_name", "")).strip() or str(record.get("updated_at", "")).strip() or now_text()
        }
        return repaired

    def _build_alarm_log_id(self, record):
        seed = "|".join([
            str(record.get("alarm_key", "")).strip(),
            str(record.get("alarm_time", "")).strip(),
            str(record.get("alarm_status", "")).strip(),
            str(record.get("metric_value", "")).strip()
        ])
        return uuid.uuid5(uuid.NAMESPACE_URL, seed or uuid.uuid4().hex).hex

    def _looks_like_log_id(self, value):
        text = str(value or "").strip()
        return len(text) == 32 and all(ch in "0123456789abcdef" for ch in text.lower())

    def _looks_like_datetime(self, value):
        text = str(value or "").strip()
        if len(text) < 19:
            return False
        try:
            datetime.strptime(text[:19], "%Y-%m-%d %H:%M:%S")
            return True
        except ValueError:
            return False

    def _write_alarm_log_records(self, records):
        normalized = []
        for record in records:
            normalized.append({
                "log_id": str(record.get("log_id", "")).strip() or uuid.uuid4().hex,
                "alarm_key": str(record.get("alarm_key", "")).strip(),
                "alarm_id": str(record.get("alarm_id", "")).strip(),
                "rule_key": str(record.get("rule_key", "")).strip(),
                "rule_name": str(record.get("rule_name", "")).strip(),
                "alarm_time": str(record.get("alarm_time", "")).strip(),
                "alarm_status": str(record.get("alarm_status", "")).strip(),
                "metric_value": str(record.get("metric_value", "")).strip(),
                "device_record_id": str(record.get("device_record_id", "")).strip(),
                "device_name": str(record.get("device_name", "")).strip(),
                "process_status": str(record.get("process_status", "")).strip() or "未处理",
                "processed_at": str(record.get("processed_at", "")).strip(),
                "process_note": str(record.get("process_note", "")).strip(),
                "process_records": json.dumps(record.get("process_records") or [], ensure_ascii=False),
                "updated_at": str(record.get("updated_at", "")).strip() or now_text()
            })
        self._write_sheet_records(ALARM_LOG_SHEET, ALARM_LOG_HEADERS, normalized)

    def _read_alarm_state_records(self):
        records = self._read_sheet_records(ALARM_STATE_SHEET, ALARM_STATE_HEADERS)
        for record in records:
            record["is_active"] = safe_int(record.get("is_active"), 0)
        return records

    def _write_alarm_state_records(self, records):
        normalized = []
        for record in records:
            normalized.append({
                "alarm_key": str(record.get("alarm_key", "")).strip(),
                "device_record_id": str(record.get("device_record_id", "")).strip(),
                "rule_key": str(record.get("rule_key", "")).strip(),
                "is_active": 1 if safe_int(record.get("is_active"), 0) else 0,
                "last_value": str(record.get("last_value", "")).strip(),
                "updated_at": str(record.get("updated_at", "")).strip() or now_text()
            })
        self._write_sheet_records(ALARM_STATE_SHEET, ALARM_STATE_HEADERS, normalized)

    def _mysql_settings(self):
        return self._load_config().get("mysql", {})

    def _mysql_is_ready(self):
        settings = self._mysql_settings()
        required = [
            settings.get("host"),
            settings.get("user"),
            settings.get("password"),
            settings.get("database")
        ]
        if not settings.get("enabled"):
            return False
        return all(required) and "YOUR_MYSQL" not in "".join(str(item) for item in required)

    def _connect_mysql(self):
        if not self._mysql_is_ready():
            return None
        settings = self._mysql_settings()
        return pymysql.connect(
            host=settings["host"],
            port=safe_int(settings.get("port"), 3306),
            user=settings["user"],
            password=settings["password"],
            database=settings["database"],
            charset=settings.get("charset", "utf8mb4"),
            connect_timeout=safe_int(settings.get("connect_timeout"), 30),
            read_timeout=safe_int(settings.get("connect_timeout"), 30),
            write_timeout=safe_int(settings.get("connect_timeout"), 30),
            cursorclass=DictCursor,
            autocommit=False
        )

    def _ensure_mysql_table(self, connection):
        table = self._mysql_settings().get("table", "onenet_devices")
        sql = f"""
        CREATE TABLE IF NOT EXISTS `{table}` (
            `record_id` VARCHAR(64) NOT NULL,
            `display_name` VARCHAR(255) NOT NULL,
            `onenet_device_name` VARCHAR(255) NOT NULL,
            `product_id` VARCHAR(128) NOT NULL,
            `user_id` VARCHAR(128) NOT NULL,
            `access_key` TEXT NOT NULL,
            `auth_version` VARCHAR(32) NOT NULL,
            `device_secret` TEXT NULL,
            `device_id` VARCHAR(128) NULL,
            `ip` VARCHAR(128) NULL,
            `start_success_count` INT NOT NULL DEFAULT 0,
            `slot_index` INT NOT NULL DEFAULT 0,
            `notes` TEXT NULL,
            `created_at` VARCHAR(32) NOT NULL,
            `updated_at` VARCHAR(32) NOT NULL,
            PRIMARY KEY (`record_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        with connection.cursor() as cursor:
            cursor.execute(sql)
            cursor.execute(f"SHOW COLUMNS FROM `{table}` LIKE 'slot_index'")
            if cursor.fetchone() is None:
                cursor.execute(f"ALTER TABLE `{table}` ADD COLUMN `slot_index` INT NOT NULL DEFAULT 0 AFTER `start_success_count`")
        connection.commit()

    def _sync_mysql_from_excel(self, connection):
        self._ensure_mysql_table(connection)
        table = self._mysql_settings().get("table", "onenet_devices")
        excel_records = [self._normalize_record(item, allow_missing=False) for item in self._read_excel_records()]
        excel_ids = [item["record_id"] for item in excel_records]
        with connection.cursor() as cursor:
            if excel_ids:
                placeholders = ",".join(["%s"] * len(excel_ids))
                cursor.execute(f"DELETE FROM `{table}` WHERE `record_id` NOT IN ({placeholders})", excel_ids)
            else:
                cursor.execute(f"DELETE FROM `{table}`")

            upsert_sql = f"""
            INSERT INTO `{table}` (
                `record_id`, `display_name`, `onenet_device_name`, `product_id`, `user_id`,
                `access_key`, `auth_version`, `device_secret`, `device_id`, `ip`,
                `start_success_count`, `slot_index`, `notes`, `created_at`, `updated_at`
            ) VALUES (
                %(record_id)s, %(display_name)s, %(onenet_device_name)s, %(product_id)s, %(user_id)s,
                %(access_key)s, %(auth_version)s, %(device_secret)s, %(device_id)s, %(ip)s,
                %(start_success_count)s, %(slot_index)s, %(notes)s, %(created_at)s, %(updated_at)s
            )
            ON DUPLICATE KEY UPDATE
                `display_name`=VALUES(`display_name`),
                `onenet_device_name`=VALUES(`onenet_device_name`),
                `product_id`=VALUES(`product_id`),
                `user_id`=VALUES(`user_id`),
                `access_key`=VALUES(`access_key`),
                `auth_version`=VALUES(`auth_version`),
                `device_secret`=VALUES(`device_secret`),
                `device_id`=VALUES(`device_id`),
                `ip`=VALUES(`ip`),
                `start_success_count`=VALUES(`start_success_count`),
                `slot_index`=VALUES(`slot_index`),
                `notes`=VALUES(`notes`),
                `created_at`=VALUES(`created_at`),
                `updated_at`=VALUES(`updated_at`)
            """
            for record in excel_records:
                cursor.execute(upsert_sql, record)
        connection.commit()

    def _read_mysql_records(self, connection):
        self._ensure_mysql_table(connection)
        table = self._mysql_settings().get("table", "onenet_devices")
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM `{table}` ORDER BY `created_at` ASC")
            rows = cursor.fetchall()
        for row in rows:
            row["start_success_count"] = safe_int(row.get("start_success_count"), 0)
            row["slot_index"] = normalize_slot_index(row.get("slot_index"))
        return rows

    def _storage_status(self, mysql_connected, force_excel=False):
        if force_excel:
            return {
                "mode": "excel",
                "toast_color": "blue",
                "toast_text": "已切换本地数据模式",
                "mysql_connected": mysql_connected
            }
        if mysql_connected:
            return {
                "mode": "mysql",
                "toast_color": "green",
                "toast_text": "MySQL连接成功，已同步本地设备库",
                "mysql_connected": True
            }
        return {
            "mode": "excel",
            "toast_color": "blue",
            "toast_text": "当前使用本地数据模式",
            "mysql_connected": False
        }

    def _normalize_record(self, payload, allow_missing=True):
        stamp = now_text()
        record = {
            "record_id": (payload.get("record_id") or "").strip() if isinstance(payload.get("record_id"), str) else payload.get("record_id"),
            "display_name": str(payload.get("display_name", "")).strip(),
            "onenet_device_name": str(payload.get("onenet_device_name", "")).strip(),
            "product_id": str(payload.get("product_id", "")).strip(),
            "user_id": str(payload.get("user_id", "")).strip(),
            "access_key": str(payload.get("access_key", "")).strip(),
            "auth_version": str(payload.get("auth_version", "")).strip() or "2020-05-29",
            "device_secret": str(payload.get("device_secret", "")).strip(),
            "device_id": str(payload.get("device_id", "")).strip(),
            "ip": str(payload.get("ip", "")).strip() or "未配置",
            "start_success_count": safe_int(payload.get("start_success_count"), 0),
            "slot_index": normalize_slot_index(payload.get("slot_index")),
            "notes": str(payload.get("notes", "")).strip(),
            "created_at": str(payload.get("created_at", "")).strip() or stamp,
            "updated_at": stamp
        }
        if not record["record_id"]:
            record["record_id"] = uuid.uuid4().hex
        if not record["display_name"]:
            record["display_name"] = record["onenet_device_name"] or "未命名设备"
        if not allow_missing:
            return record
        required = ["display_name", "onenet_device_name", "product_id", "user_id", "access_key"]
        missing = [key for key in required if not record[key]]
        if missing:
            raise ValueError("缺少必填项: " + "、".join(missing))
        return record

    def _find_first_available_slot(self, records, exclude_record_id=None):
        used_slots = set()
        for item in records:
            if exclude_record_id and item.get("record_id") == exclude_record_id:
                continue
            slot_index = normalize_slot_index(item.get("slot_index"))
            if slot_index:
                used_slots.add(slot_index)
        for slot_index in range(1, 21):
            if slot_index not in used_slots:
                return slot_index
        return 0

    def _apply_slot_assignment(self, record, records, exclude_record_id=None):
        slot_index = normalize_slot_index(record.get("slot_index"))
        for item in records:
            if exclude_record_id and item.get("record_id") == exclude_record_id:
                continue
            if normalize_slot_index(item.get("slot_index")) == slot_index and slot_index:
                raise ValueError(f"RCS地图槽位 {slot_index} 已被占用")
        if not slot_index:
            slot_index = self._find_first_available_slot(records, exclude_record_id=exclude_record_id)
        if not slot_index:
            raise ValueError("RCS地图最多只能放置20个设备，请先删减或调整现有设备")
        record["slot_index"] = slot_index
        return record

    def _query_onenet(self, record, path, params):
        authorization = self._create_authorization(record["access_key"], record["user_id"], record["auth_version"])
        response = requests.get(
            "https://iot-api.heclouds.com" + path,
            params=params,
            headers={"authorization": authorization},
            timeout=15
        )
        response.raise_for_status()
        data = response.json()
        if data.get("code") != 0:
            raise RuntimeError(data.get("msg") or "OneNET请求失败")
        return data

    def _create_authorization(self, access_key, user_id, version):
        method = "sha1"
        res = f"userid/{user_id}"
        et = str(int(datetime.now().timestamp()) + 3600)
        sign_text = et + "\n" + method + "\n" + res + "\n" + version
        try:
            key = base64.b64decode(access_key)
        except Exception:
            key = access_key.encode("utf-8")
        sign = base64.b64encode(hmac.new(key, sign_text.encode("utf-8"), hashlib.sha1).digest()).decode("utf-8")
        return "version={}&res={}&et={}&method={}&sign={}".format(
            version,
            quote(res, safe=""),
            et,
            method,
            quote(sign, safe="")
        )

    def _decorate_device(self, record):
        device = {
            "record_id": record["record_id"],
            "id": record.get("device_id") or record["record_id"],
            "device_name": record["onenet_device_name"],
            "name": record["display_name"],
            "ip": record.get("ip") or "未配置",
            "starts": safe_int(record.get("start_success_count"), 0),
            "slot_index": normalize_slot_index(record.get("slot_index")),
            "online": False,
            "status_text": "unknown",
            "created_at": record.get("created_at", ""),
            "raw": copy.deepcopy(record)
        }
        try:
            detail = self._query_onenet(record, "/device/detail", {
                "product_id": record["product_id"],
                "device_name": record["onenet_device_name"]
            })
            remote = detail.get("data") or {}
            if remote.get("did"):
                device["id"] = str(remote["did"])
            device["online"] = remote.get("status") == 1
            device["status_text"] = "online" if device["online"] else "offline"
        except Exception:
            device["online"] = False
            device["status_text"] = "unknown"
        return device

    def _refresh_status_statistics(self, force_excel=False):
        records, status = self.list_records(force_excel=force_excel, decorate=True)
        now_stamp = now_text()
        today = datetime.now().strftime("%Y-%m-%d")
        status_rows = self._read_device_status_records()
        history_rows = self._read_status_history_records()
        status_map = {item.get("record_id"): item for item in status_rows}
        history_map = {item.get("date"): item for item in history_rows if item.get("date")}
        next_status_rows = []
        online_count = 0
        offline_count = 0
        unknown_count = 0
        status_changed = len(status_rows) != len(records)

        for device in records:
            status_text = device.get("status_text") or "unknown"
            if status_text == "online":
                online_count += 1
            elif status_text == "offline":
                offline_count += 1
            else:
                unknown_count += 1
            previous = status_map.get(device["record_id"], {})
            last_seen_at = previous.get("last_seen_at") or now_stamp
            if previous.get("status") != status_text:
                last_seen_at = now_stamp
                status_changed = True
            next_status_rows.append({
                "record_id": device["record_id"],
                "device_name": device.get("device_name", ""),
                "display_name": device.get("name", ""),
                "status": status_text,
                "online": 1 if status_text == "online" else 0,
                "last_seen_at": last_seen_at,
                "updated_at": now_stamp
            })

        if status_changed:
            self._write_device_status_records(next_status_rows)

        total_devices = len(records)
        today_row = history_map.get(today, {})
        next_today_row = {
            "date": today,
            "online_count": online_count,
            "offline_count": offline_count + unknown_count,
            "unknown_count": unknown_count,
            "total_devices": total_devices,
            "updated_at": now_stamp
        }
        if (
            safe_int(today_row.get("online_count"), -1) != next_today_row["online_count"]
            or safe_int(today_row.get("offline_count"), -1) != next_today_row["offline_count"]
            or safe_int(today_row.get("unknown_count"), -1) != next_today_row["unknown_count"]
            or safe_int(today_row.get("total_devices"), -1) != next_today_row["total_devices"]
        ):
            history_map[today] = next_today_row
            history_list = sorted(history_map.values(), key=lambda item: item.get("date", ""))
            self._write_status_history_records(history_list)

        return {
            "records": records,
            "storage": status,
            "online_count": online_count,
            "offline_count": offline_count + unknown_count,
            "unknown_count": unknown_count,
            "total_devices": total_devices
        }

    def _find_record(self, record_id, force_excel=False):
        records, _ = self.list_records(force_excel=force_excel, decorate=False)
        for record in records:
            if record["record_id"] == record_id:
                return record
        raise KeyError("未找到指定设备")

    def list_records(self, force_excel=False, decorate=True):
        with self.lock:
            self._ensure_seed_data()
            if force_excel:
                records = self._read_excel_records()
                status = self._storage_status(mysql_connected=self._mysql_is_ready(), force_excel=True)
            else:
                try:
                    connection = self._connect_mysql()
                except Exception:
                    connection = None
                if connection:
                    try:
                        self._sync_mysql_from_excel(connection)
                        records = self._read_mysql_records(connection)
                        status = self._storage_status(mysql_connected=True, force_excel=False)
                    finally:
                        connection.close()
                else:
                    records = self._read_excel_records()
                    status = self._storage_status(mysql_connected=False, force_excel=False)

        if decorate:
            with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                decorated_records = list(executor.map(self._decorate_device, records))
            return decorated_records, status
        return records, status

    def add_device(self, payload):
        record = self._normalize_record(payload, allow_missing=True)
        with self.lock:
            records = self._read_excel_records()
            records = [item for item in records if item["record_id"] != record["record_id"]]
            record = self._apply_slot_assignment(record, records)
            records.append(record)
            self._write_excel_records(records)

            mysql_connected = False
            try:
                connection = self._connect_mysql()
            except Exception:
                connection = None
            if connection:
                try:
                    self._sync_mysql_from_excel(connection)
                    mysql_connected = True
                finally:
                    connection.close()
        return self._decorate_device(record), self._storage_status(mysql_connected, force_excel=False)

    def update_device(self, record_id, payload):
        with self.lock:
            records = self._read_excel_records()
            target = None
            next_records = []
            for item in records:
                if item["record_id"] == record_id:
                    target = item
                else:
                    next_records.append(item)
            if target is None:
                raise KeyError("未找到需要修改的设备")
            merged = dict(target)
            merged.update(payload or {})
            merged["record_id"] = record_id
            merged["created_at"] = target.get("created_at", "")
            record = self._normalize_record(merged, allow_missing=True)
            record = self._apply_slot_assignment(record, records, exclude_record_id=record_id)
            next_records.append(record)
            self._write_excel_records(next_records)

            mysql_connected = False
            try:
                connection = self._connect_mysql()
            except Exception:
                connection = None
            if connection:
                try:
                    self._sync_mysql_from_excel(connection)
                    mysql_connected = True
                finally:
                    connection.close()
        return self._decorate_device(record), self._storage_status(mysql_connected, force_excel=False)

    def delete_device(self, record_id):
        with self.lock:
            records = self._read_excel_records()
            next_records = [item for item in records if item["record_id"] != record_id]
            if len(next_records) == len(records):
                raise KeyError("未找到需要删减的设备")
            self._write_excel_records(next_records)

            mysql_connected = False
            try:
                connection = self._connect_mysql()
            except Exception:
                connection = None
            if connection:
                try:
                    self._sync_mysql_from_excel(connection)
                    mysql_connected = True
                finally:
                    connection.close()
        return self._storage_status(mysql_connected, force_excel=False)

    def get_metrics(self, force_excel=False):
        snapshot = self._refresh_status_statistics(force_excel=force_excel)
        records = snapshot["records"]
        status = snapshot["storage"]
        current_month = datetime.now().strftime("%Y-%m")
        new_devices = 0
        for d in records:
            if d.get("created_at", "").startswith(current_month):
                new_devices += 1
        alarm_snapshot = self._refresh_local_alarms(force_excel=force_excel, limit=300)
        error_count = self._count_recent_alarm_logs(alarm_snapshot.get("all_items") or [], days=5)

        return {
            "error_count": error_count,
            "new_devices_this_month": new_devices,
            "online_devices": snapshot["online_count"],
            "total_devices": snapshot["total_devices"],
            "storage": status
        }

    def get_status_trend(self, force_excel=False):
        snapshot = self._refresh_status_statistics(force_excel=force_excel)
        history_map = {
            item.get("date"): item
            for item in self._read_status_history_records()
            if item.get("date")
        }
        today = datetime.now().date()
        items = []
        for offset in range(4, -1, -1):
            current_date = today - timedelta(days=offset)
            date_key = current_date.strftime("%Y-%m-%d")
            history_row = history_map.get(date_key, {})
            items.append({
                "date": date_key,
                "label": current_date.strftime("%m-%d"),
                "online_count": safe_int(history_row.get("online_count"), 0),
                "offline_count": safe_int(history_row.get("offline_count"), 0),
                "unknown_count": safe_int(history_row.get("unknown_count"), 0),
                "total_devices": safe_int(history_row.get("total_devices"), snapshot["total_devices"] if offset == 0 else 0)
            })
        return {
            "items": items,
            "today": {
                "online_count": snapshot["online_count"],
                "offline_count": snapshot["offline_count"],
                "unknown_count": snapshot["unknown_count"],
                "total_devices": snapshot["total_devices"]
            },
            "storage": snapshot["storage"]
        }

    def get_alarm_config(self):
        with self.lock:
            self._ensure_seed_data()
        config_rows = self._read_alarm_config_records()
        config_rows.sort(key=lambda item: item.get("rule_key", ""))
        return {
            "rules": config_rows
        }

    def save_alarm_config(self, payload):
        payload_rules = payload.get("rules") if isinstance(payload, dict) else None
        if not isinstance(payload_rules, list):
            raise ValueError("缺少 rules 配置")
        current_rows = {
            item.get("rule_key"): item
            for item in self._read_alarm_config_records()
        }
        next_rows = []
        for default_rule in DEFAULT_ALARM_CONFIG:
            rule_key = default_rule["rule_key"]
            incoming = None
            for item in payload_rules:
                if isinstance(item, dict) and item.get("rule_key") == rule_key:
                    incoming = item
                    break
            base_row = current_rows.get(rule_key, default_rule)
            next_rows.append({
                "rule_key": rule_key,
                "rule_name": default_rule["rule_name"],
                "threshold_value": safe_float((incoming or {}).get("threshold_value", base_row.get("threshold_value"))),
                "alarm_level": default_rule["alarm_level"],
                "enabled": 1 if safe_int((incoming or {}).get("enabled", base_row.get("enabled", 1)), 1) else 0,
                "updated_at": now_text()
            })
        self._write_alarm_config_records(next_rows)
        return {"rules": next_rows}

    def get_alarm_list(self, force_excel=False, limit=20):
        return self._refresh_local_alarms(force_excel=force_excel, limit=limit)

    def _refresh_local_alarms(self, force_excel=False, limit=20):
        with self.lock:
            self._ensure_seed_data()
        config_rows = self._read_alarm_config_records()
        config_map = {item["rule_key"]: item for item in config_rows}
        state_rows = self._read_alarm_state_records()
        state_map = {item["alarm_key"]: item for item in state_rows}
        log_rows = self._read_alarm_log_records()
        records, storage = self.list_records(force_excel=force_excel, decorate=False)
        now_stamp = now_text()
        next_state_rows = []
        new_alarm_rows = []
        active_alarm_rows = []

        for record in records:
            device_name = record.get("display_name") or record.get("onenet_device_name") or "未命名设备"
            try:
                prop = self.get_device_properties(record["record_id"], force_excel=force_excel)
            except Exception:
                prop = {
                    "temp": "",
                    "humi": "",
                    "smoke": "",
                    "status": "unknown"
                }

            evaluations = self._build_alarm_evaluations(record, device_name, prop, config_map, now_stamp)
            for item in evaluations:
                alarm_key = item["alarm_key"]
                previous = state_map.get(alarm_key, {})
                was_active = safe_int(previous.get("is_active"), 0) == 1
                is_active = item["is_active"]
                next_state_rows.append({
                    "alarm_key": alarm_key,
                    "device_record_id": record["record_id"],
                    "rule_key": item["rule_key"],
                    "is_active": 1 if is_active else 0,
                    "last_value": item["metric_value"],
                    "updated_at": now_stamp
                })
                if is_active:
                    active_alarm_rows.append(item)
                    if not was_active:
                        log_item = self._build_alarm_log_item(item)
                        log_rows.append(log_item)
                        new_alarm_rows.append(log_item)

        self._write_alarm_state_records(next_state_rows)
        log_rows.sort(key=lambda item: str(item.get("alarm_time", "")), reverse=True)
        self._write_alarm_log_records(log_rows[:300])
        return {
            "items": log_rows[:safe_int(limit, 20)],
            "all_items": log_rows,
            "new_alarm_items": new_alarm_rows,
            "active_alarm_items": active_alarm_rows,
            "config": {"rules": config_rows},
            "storage": storage
        }

    def _build_alarm_log_item(self, item):
        log_item = dict(item)
        log_item["log_id"] = uuid.uuid4().hex
        log_item["process_status"] = "未处理"
        log_item["processed_at"] = ""
        log_item["process_note"] = ""
        log_item["process_records"] = []
        return log_item

    def get_alarm_detail(self, log_id):
        logs = self._read_alarm_log_records()
        target = next((item for item in logs if item.get("log_id") == log_id), None)
        if target is None:
            raise KeyError("未找到指定告警")
        history = [item for item in logs if item.get("alarm_key") == target.get("alarm_key")]
        history.sort(key=lambda item: str(item.get("alarm_time", "")), reverse=True)
        return {
            "current": target,
            "history": history
        }

    def mark_alarm_processed(self, log_id):
        logs = self._read_alarm_log_records()
        target = None
        now_stamp = now_text()
        for item in logs:
            if item.get("log_id") != log_id:
                continue
            target = item
            if item.get("process_status") != "已处理":
                item["process_status"] = "已处理"
                item["processed_at"] = now_stamp
                item["process_note"] = "点击告警日志后自动处理"
                records = item.get("process_records") or []
                records.append({
                    "time": now_stamp,
                    "action": "已处理",
                    "operator": "root",
                    "note": item["process_note"]
                })
                item["process_records"] = records
                item["updated_at"] = now_stamp
            break
        if target is None:
            raise KeyError("未找到指定告警")
        self._write_alarm_log_records(logs)
        return self.get_alarm_detail(log_id)

    def _count_recent_alarm_logs(self, alarm_rows, days=5):
        start_date = (datetime.now() - timedelta(days=max(days - 1, 0))).date()
        count = 0
        for row in alarm_rows:
            alarm_time = str(row.get("alarm_time", "")).strip()
            try:
                alarm_date = datetime.strptime(alarm_time[:19], "%Y-%m-%d %H:%M:%S").date()
            except ValueError:
                continue
            if alarm_date >= start_date:
                count += 1
        return count

    def _build_alarm_evaluations(self, record, device_name, prop, config_map, alarm_time):
        rows = []
        temp_value = safe_float(prop.get("temp"), 0)
        humi_value = safe_float(prop.get("humi"), 0)
        smoke_value = safe_float(prop.get("smoke"), 0)
        status_value = str(prop.get("status", "unknown")).lower()
        status_text = "在线" if status_value == "true" else ("离线" if status_value == "false" else "未知")

        metric_specs = [
            ("temp", temp_value, lambda cfg: temp_value > cfg["threshold_value"], str(temp_value)),
            ("humi", humi_value, lambda cfg: humi_value > cfg["threshold_value"], str(humi_value)),
            ("smoke", smoke_value, lambda cfg: smoke_value > cfg["threshold_value"], str(smoke_value)),
            ("status", status_text, lambda cfg: status_value != "true", status_text)
        ]

        for rule_key, _, rule_match, display_value in metric_specs:
            config = config_map.get(rule_key)
            if not config or not safe_int(config.get("enabled"), 0):
                continue
            is_active = bool(rule_match(config))
            rows.append({
                "alarm_key": record["record_id"] + ":" + rule_key,
                "alarm_id": device_name,
                "rule_key": rule_key,
                "rule_name": config.get("rule_name", rule_key),
                "alarm_time": alarm_time,
                "alarm_status": config.get("alarm_level", "提醒"),
                "metric_value": str(display_value),
                "device_record_id": record["record_id"],
                "device_name": device_name,
                "updated_at": alarm_time,
                "is_active": is_active
            })
        return rows

    def get_device_properties(self, record_id, force_excel=False):
        record = self._find_record(record_id, force_excel=force_excel)
        property_data = self._query_onenet(record, "/thingmodel/query-device-property", {
            "product_id": record["product_id"],
            "device_name": record["onenet_device_name"]
        })
        property_map = {}
        for item in property_data.get("data") or []:
            property_map[item.get("identifier")] = item.get("value")

        decorated = self._decorate_device(record)
        return {
            "record_id": record["record_id"],
            "device_id": decorated["id"],
            "display_name": record["display_name"],
            "device_name": record["onenet_device_name"],
            "ip": record.get("ip") or "未配置",
            "temp": property_map.get("temp", ""),
            "humi": property_map.get("humi", ""),
            "smoke": property_map.get("smoke", ""),
            "status": property_map.get("status", "false")
        }

    def set_device_properties(self, record_id, payload, force_excel=False):
        record = self._find_record(record_id, force_excel=force_excel)
        params = {}
        for key in ("temp", "humi", "smoke", "status"):
            if key not in payload:
                continue
            value = payload.get(key)
            if key == "status":
                if isinstance(value, str):
                    value = value.strip().lower() in ("1", "true", "on", "online")
                else:
                    value = bool(value)
            else:
                value = safe_float(value, None)
                if value is None:
                    continue
            params[key] = value
        if not params:
            raise ValueError("缺少可控制的属性")

        authorization = self._create_authorization(record["access_key"], record["user_id"], record["auth_version"])
        response = requests.post(
            "https://iot-api.heclouds.com/thingmodel/set-device-property",
            json={
                "product_id": record["product_id"],
                "device_name": record["onenet_device_name"],
                "params": params
            },
            headers={
                "authorization": authorization,
                "Content-Type": "application/json"
            },
            timeout=15
        )
        response.raise_for_status()
        data = response.json()
        if data.get("code") != 0:
            raise RuntimeError(data.get("msg") or "OneNET属性下发失败")
        result = self.get_device_properties(record_id, force_excel=force_excel)
        result["sent_params"] = params
        return result


class DeviceRequestHandler(BaseHTTPRequestHandler):
    repository = DeviceRepository()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/login.html":
            self._serve_static(parsed.path)
            return
        if not self._is_authenticated():
            if parsed.path.startswith("/api/"):
                self._send_json(HTTPStatus.UNAUTHORIZED, {"code": 401, "msg": "未登录"})
            else:
                self._redirect_to_login()
            return
        if parsed.path == "/api/devices":
            self._handle_list_devices(parsed)
            return
        if parsed.path == "/api/device-properties":
            self._handle_device_properties(parsed)
            return
        if parsed.path == "/api/metrics":
            self._handle_metrics(parsed)
            return
        if parsed.path == "/api/status-trend":
            self._handle_status_trend(parsed)
            return
        if parsed.path == "/api/alarms":
            self._handle_alarms(parsed)
            return
        if parsed.path == "/api/alarm-config":
            self._handle_alarm_config_get()
            return
        if parsed.path == "/api/alarm-detail":
            self._handle_alarm_detail(parsed)
            return
        if parsed.path == "/api/app-demo-device":
            self._handle_app_demo_device(parsed)
            return
        if parsed.path == "/api/weather":
            self._handle_weather(parsed)
            return
        self._serve_static(parsed.path)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/login":
            self._handle_login()
            return
        if not self._is_authenticated():
            self._send_json(HTTPStatus.UNAUTHORIZED, {"code": 401, "msg": "未登录"})
            return
        if parsed.path == "/api/devices":
            self._handle_add_device()
            return
        if parsed.path == "/api/alarm-config":
            self._handle_alarm_config_save()
            return
        if parsed.path == "/api/device-control":
            self._handle_device_control(parsed)
            return
        if parsed.path == "/api/alarm-process":
            self._handle_alarm_process()
            return
        self._send_json(HTTPStatus.NOT_FOUND, {"code": 404, "msg": "接口不存在"})

    def do_PUT(self):
        parsed = urlparse(self.path)
        if not self._is_authenticated():
            self._send_json(HTTPStatus.UNAUTHORIZED, {"code": 401, "msg": "未登录"})
            return
        if parsed.path.startswith("/api/devices/"):
            self._handle_update_device(parsed.path.rsplit("/", 1)[-1])
            return
        self._send_json(HTTPStatus.NOT_FOUND, {"code": 404, "msg": "接口不存在"})

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if not self._is_authenticated():
            self._send_json(HTTPStatus.UNAUTHORIZED, {"code": 401, "msg": "未登录"})
            return
        if parsed.path.startswith("/api/devices/"):
            self._handle_delete_device(parsed.path.rsplit("/", 1)[-1])
            return
        self._send_json(HTTPStatus.NOT_FOUND, {"code": 404, "msg": "接口不存在"})

    def _handle_list_devices(self, parsed):
        query = parse_qs(parsed.query)
        force_excel = query.get("force_excel", ["0"])[0] == "1"
        devices, status = self.repository.list_records(force_excel=force_excel, decorate=True)
        self._send_json(HTTPStatus.OK, {
            "code": 0,
            "msg": "succ",
            "data": {
                "devices": devices,
                "storage": status
            }
        })

    def _handle_device_properties(self, parsed):
        query = parse_qs(parsed.query)
        record_id = query.get("record_id", [""])[0]
        force_excel = query.get("force_excel", ["0"])[0] == "1"
        if not record_id:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": "缺少 record_id"})
            return
        try:
            data = self.repository.get_device_properties(record_id, force_excel=force_excel)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_metrics(self, parsed):
        query = parse_qs(parsed.query)
        force_excel = query.get("force_excel", ["0"])[0] == "1"
        try:
            data = self.repository.get_metrics(force_excel=force_excel)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_status_trend(self, parsed):
        query = parse_qs(parsed.query)
        force_excel = query.get("force_excel", ["0"])[0] == "1"
        try:
            data = self.repository.get_status_trend(force_excel=force_excel)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_alarms(self, parsed):
        query = parse_qs(parsed.query)
        force_excel = query.get("force_excel", ["0"])[0] == "1"
        limit = safe_int(query.get("limit", ["20"])[0], 20)
        try:
            data = self.repository.get_alarm_list(force_excel=force_excel, limit=limit)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_alarm_config_get(self):
        try:
            data = self.repository.get_alarm_config()
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_alarm_detail(self, parsed):
        query = parse_qs(parsed.query)
        log_id = query.get("log_id", [""])[0]
        if not log_id:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": "缺少 log_id"})
            return
        try:
            data = self.repository.get_alarm_detail(log_id)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data})
        except KeyError as error:
            self._send_json(HTTPStatus.NOT_FOUND, {"code": 404, "msg": str(error)})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_alarm_config_save(self):
        body = self._read_json_body()
        try:
            data = self.repository.save_alarm_config(body)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "阈值已保存", "data": data})
        except ValueError as error:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": str(error)})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_app_demo_device(self, parsed):
        query = parse_qs(parsed.query)
        force_excel = query.get("force_excel", ["0"])[0] == "1"
        record_id = query.get("record_id", [""])[0]
        try:
            devices, storage = self.repository.list_records(force_excel=force_excel, decorate=True)
            current = None
            if record_id:
                current = next((item for item in devices if item["record_id"] == record_id), None)
            if current is None and devices:
                current = devices[0]
            if current is None:
                self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": {"device": None, "storage": storage}})
                return
            data = self.repository.get_device_properties(current["record_id"], force_excel=force_excel)
            data["storage"] = storage
            data["online"] = current.get("online", False)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_device_control(self, parsed):
        query = parse_qs(parsed.query)
        force_excel = query.get("force_excel", ["0"])[0] == "1"
        body = self._read_json_body()
        record_id = str(body.get("record_id", "")).strip()
        if not record_id:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": "缺少 record_id"})
            return
        try:
            data = self.repository.set_device_properties(record_id, body, force_excel=force_excel)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "控制下发成功", "data": data})
        except ValueError as error:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": str(error)})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_weather(self, parsed):
        query = parse_qs(parsed.query)
        city_code = query.get("city_code", ["101010100"])[0]
        try:
            url = f"http://t.weather.itboy.net/api/weather/city/{city_code}"
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            data = response.json()
            if data.get("status") != 200:
                raise RuntimeError(data.get("message") or "获取天气失败")
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "succ", "data": data.get("data")})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_add_device(self):
        body = self._read_json_body()
        try:
            device, status = self.repository.add_device(body)
            self._send_json(HTTPStatus.OK, {
                "code": 0,
                "msg": "设备已保存",
                "data": {
                    "device": device,
                    "storage": status
                }
            })
        except ValueError as error:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": str(error)})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_update_device(self, record_id):
        body = self._read_json_body()
        try:
            device, status = self.repository.update_device(record_id, body)
            self._send_json(HTTPStatus.OK, {
                "code": 0,
                "msg": "设备已更新",
                "data": {
                    "device": device,
                    "storage": status
                }
            })
        except KeyError as error:
            self._send_json(HTTPStatus.NOT_FOUND, {"code": 404, "msg": str(error)})
        except ValueError as error:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": str(error)})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_login(self):
        body = self._read_json_body()
        username = str(body.get("username", "")).strip()
        password = str(body.get("password", "")).strip()
        if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
            self._send_json(
                HTTPStatus.OK,
                {"code": 0, "msg": "登录成功"},
                extra_headers=[
                    ("Set-Cookie", f"{AUTH_COOKIE_NAME}={AUTH_COOKIE_VALUE}; Path=/; HttpOnly; SameSite=Lax")
                ]
            )
            return
        self._send_json(HTTPStatus.UNAUTHORIZED, {"code": 401, "msg": "登陆失败"})

    def _handle_delete_device(self, record_id):
        try:
            status = self.repository.delete_device(record_id)
            self._send_json(HTTPStatus.OK, {
                "code": 0,
                "msg": "设备已删减",
                "data": {
                    "storage": status
                }
            })
        except KeyError as error:
            self._send_json(HTTPStatus.NOT_FOUND, {"code": 404, "msg": str(error)})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _handle_alarm_process(self):
        body = self._read_json_body()
        log_id = str(body.get("log_id", "")).strip()
        if not log_id:
            self._send_json(HTTPStatus.BAD_REQUEST, {"code": 400, "msg": "缺少 log_id"})
            return
        try:
            data = self.repository.mark_alarm_processed(log_id)
            self._send_json(HTTPStatus.OK, {"code": 0, "msg": "告警已处理", "data": data})
        except KeyError as error:
            self._send_json(HTTPStatus.NOT_FOUND, {"code": 404, "msg": str(error)})
        except Exception as error:
            self._send_json(HTTPStatus.BAD_GATEWAY, {"code": 500, "msg": str(error)})

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length) if length > 0 else b"{}"
        return json.loads(raw.decode("utf-8") or "{}")

    def _serve_static(self, raw_path):
        relative = raw_path or "/"
        if relative == "/":
            relative = "/index.html"

        root_file_path = (ROOT_DIR / relative.lstrip("/")).resolve()
        dist_file_path = (FRONTEND_DIST_DIR / relative.lstrip("/")).resolve()

        public_dirs = [
            ROOT_DIR / "css", ROOT_DIR / "js", ROOT_DIR / "img",
            FRONTEND_DIST_DIR / "assets", FRONTEND_DIST_DIR / "css",
            FRONTEND_DIST_DIR / "js", FRONTEND_DIST_DIR / "img"
        ]
        public_files = {
            ROOT_DIR / "index.html",
            ROOT_DIR / "login.html",
            ROOT_DIR / "app_demo.html",
            FRONTEND_DIST_DIR / "index.html",
            FRONTEND_DIST_DIR / "login.html",
            FRONTEND_DIST_DIR / "app_demo.html"
        }

        file_path = None
        if FRONTEND_DIST_DIR.exists() and dist_file_path.exists() and dist_file_path.is_file():
            file_path = dist_file_path
        elif root_file_path.exists() and root_file_path.is_file():
            file_path = root_file_path

        if file_path is None:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        is_public_file = file_path in public_files
        is_public_dir_file = any(file_path.is_relative_to(item) for item in public_dirs if item.exists())
        if not is_public_file and not is_public_dir_file:
            self.send_error(HTTPStatus.FORBIDDEN)
            return

        mime_type, _ = mimetypes.guess_type(str(file_path))
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mime_type or "application/octet-stream")
        self.send_header("Content-Length", str(file_path.stat().st_size))
        self.end_headers()
        with open(file_path, "rb") as file_obj:
            self.wfile.write(file_obj.read())

    def _is_authenticated(self):
        cookie_text = self.headers.get("Cookie", "")
        if not cookie_text:
            return False
        cookie_items = {}
        for chunk in cookie_text.split(";"):
            if "=" not in chunk:
                continue
            key, value = chunk.split("=", 1)
            cookie_items[key.strip()] = value.strip()
        return cookie_items.get(AUTH_COOKIE_NAME) == AUTH_COOKIE_VALUE

    def _redirect_to_login(self):
        self.send_response(HTTPStatus.FOUND)
        self.send_header("Location", "/login.html")
        self.end_headers()

    def _send_json(self, status, payload, extra_headers=None):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        if extra_headers:
            for key, value in extra_headers:
                self.send_header(key, value)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format_string, *args):
        return


def main():
    config = copy.deepcopy(DEFAULT_CONFIG)
    if CONFIG_PATH.exists():
        try:
            deep_merge(config, json.loads(CONFIG_PATH.read_text(encoding="utf-8")))
        except json.JSONDecodeError:
            pass
    host = config.get("server", {}).get("host", "127.0.0.1")
    port = safe_int(config.get("server", {}).get("port"), 8000)
    server = ThreadingHTTPServer((host, port), DeviceRequestHandler)
    print(f"设备服务已启动: http://{host}:{port}/")
    print(f"本地Excel表: {EXCEL_PATH}")
    print(f"MySQL配置文件: {CONFIG_PATH}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
